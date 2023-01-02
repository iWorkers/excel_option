from openpyxl  import load_workbook
import random
import time
import datetime

#  3.4.4 xlwings读写 Excel
def fun3_4_4(t):
    
    wb = load_workbook('注塑产品信息清单.xlsx')
    #print(wb.sheetnames)
    ws = wb.active
    num=rand_num(get_max_row(ws))
    jt=[]
    jt=num[1]
    i=0
    data=[]
    for r in num[0]:
        rows = ws[r]
        da=[]
        da.append(t.strftime('%Y.%m.%d'))
        da.append(rows[0].value)
        da.append(jt[i])
        da.append(rows[2].value)
        da.append(creat_num(t,rows[0].value))
        i=i+1
        data.append(da)
    
    write_data(data)
    print(t.strftime('%Y.%m.%d')+"记录完成")
    
def write_data(data):
    
    wb = load_workbook('test.xlsx')
    ws = wb.active
    
    rows = data
    
    # 按行写入
    for row in rows:
        ws.append(row)
    wb.save("test.xlsx")

def creat_num(t,pinhao):
    if len(str(pinhao))==6:
        date_str = t.strftime('%y%m')+'010CX'
    else:
        date_str = t.strftime('%y%m%d')+'0CX'
    return date_str
    
def rand_num(row):
    random_nums_list1=[]
    random_nums_list2=[]
    random_nums_list=[]
    for i in range(1,random.randint(6,8)) :
        random_nums_list1.append(random.randint(1,row))
        random_nums_list2.append('A'+str(random.randint(1,20)))
    random_nums_list=[random_nums_list1,random_nums_list2]
    return(random_nums_list)

def get_max_row(sheet):
    i=sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_row = i
            break
    return real_max_row

def dayadd(t):   
    delta = datetime.timedelta(days=1)
    n_days = t + delta
    return(n_days)

if __name__=='__main__' :
    
    t1=input("开始日期:")
    t2=input("结束日期:")
    start=datetime.datetime.strptime(t1, "%Y-%m-%d")
    end=datetime.datetime.strptime(t2, "%Y-%m-%d")
    while end>start:
        fun3_4_4(start)
        start=dayadd(start)
    
    
